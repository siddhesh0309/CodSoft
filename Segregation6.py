import pandas as pd
import numpy as np
from openpyxl import load_workbook

# File paths
template_path = "template.xlsm"
mapping_path = "mapping_file.xlsx"

# Step 1: Read the template file (headers in row 4, data starts at row 9)
print("🔄 Loading template data...")
template_df = pd.read_excel(template_path, sheet_name="Account Holder", engine="openpyxl", header=3)

# Step 2: Read the mapping file
print("🔄 Loading mapping data...")
mapping_df = pd.read_excel(mapping_path)

# Step 3: Merge on 'Account Number' to bring in Occupation, GHO Code, and Status of Account
merged_df = template_df.merge(
    mapping_df[["Account Number", "Occupation", "GHO Code", "Status of Account"]],
    on="Account Number",
    how="left"
)

# Step 4: Logic for Sea Fearer Occupation adjustment
foreign_tin_cols = ["Foreign_TIN"] + [f"Foreign_TIN{i}" for i in range(2, 9)]
tin_country_cols = ["TIN_Issuing_Country"] + [f"TIN_Issuing_Country{i}" for i in range(2, 9)]

def adjust_occupation(row):
    if row["Occupation"] != "Sea Fearer":
        return row["Occupation"]
    
    has_real_tin = False
    has_dummy_or_missing_tin = True  # assume true until proven otherwise
    
    for tin_col, country_col in zip(foreign_tin_cols, tin_country_cols):
        tin = row.get(tin_col)
        country = row.get(country_col)

        if pd.notna(tin) and tin != "" and tin != "AAAAAAAAA":
            if country != "IN":
                has_real_tin = True
                has_dummy_or_missing_tin = False
        elif pd.isna(tin) or tin == "" or tin == "AAAAAAAAA":
            continue
        else:
            has_dummy_or_missing_tin = False

    if has_real_tin and row.get("SC") == "Y":
        return ""  # Blank Occupation
    elif has_dummy_or_missing_tin:
        return "Sea Fearer"
    else:
        return row["Occupation"]

# Apply Sea Fearer logic
print("⚙️ Applying Sea Fearer logic...")
merged_df["Occupation"] = merged_df.apply(adjust_occupation, axis=1)

# Step 5: Save the updated data back to the template (starting from row 9)
print("💾 Writing updated data back to template...")
wb = load_workbook(template_path, keep_vba=True)
ws = wb["Account Holder"]

# Get final headers and write them to row 4
headers = list(merged_df.columns)
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=4, column=col_idx, value=header)

# Clear existing data from row 9 onward
for row in ws.iter_rows(min_row=9, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Write updated data starting from row 9
for row_idx, row in enumerate(merged_df.itertuples(index=False), start=9):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(template_path)
print("✅ Template file updated successfully.")

# Step 6: Create segregated files by Occupation
print("📂 Creating segregated files by Occupation...")
unique_occupations = merged_df["Occupation"].dropna().unique()

for occ in unique_occupations:
    occ_df = merged_df[merged_df["Occupation"] == occ]
    filename = f"occupation_{occ.replace(' ', '_')}.xlsx"
    occ_df.to_excel(filename, index=False)
    print(f"✅ File created: {filename}")

print("🎯 All operations completed successfully.")
